import os
import numpy as np
from openpyxl import Workbook
# 读取txt文件
directory = './sample/'
sample_doc_name=os.listdir(directory)
sample_doc_name.sort(key= lambda x:int(x[6:len(x)-4]))  # 将文件名的数字提取出来进行排序
data=()
for file in sample_doc_name:
    temp_data=[]
    with open(directory+file,'r') as f:
        for line in f:
            temp=list(map(float,line.split()))
            if temp:
                temp_data+=temp
    data+=(temp_data,)
wb=Workbook()
ws=wb.active
independ_variable=np.linspace(501,7999,3750)/100
for i in range(len(independ_variable)):
    ws.cell(row=i+1,column=1,value=independ_variable[i])
for i in range(len(data)):
    temp=data[i][1::2]
    for j in range(len(temp)):
        ws.cell(row=j+1,column=i+2,value=temp[j])
# 存储的文件名可以改一下
wb.save('sample.xlsx')